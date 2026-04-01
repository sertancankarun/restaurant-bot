import os
import random
from datetime import datetime

import openpyxl
from telegram import (
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    Update,
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
)

TOKEN = os.getenv("TOKEN")
ADMIN_ID_STR = os.getenv("ADMIN_ID")

if not TOKEN:
    raise ValueError("TOKEN tanımlı değil")

if not ADMIN_ID_STR:
    raise ValueError("ADMIN_ID tanımlı değil")

ADMIN_ID = int(ADMIN_ID_STR)

FILE_NAME = "orders.xlsx"
RESTAURANT_NAME = "Getiriyosun Bot"

OPEN_HOUR = 10
CLOSE_HOUR = 23

MENU_ITEMS = {
    "pizza": 150,
    "burger": 120,
    "ayran": 30,
    "kola": 40,
}

EMOJI = {
    "pizza": "🍕",
    "burger": "🍔",
    "ayran": "🥛",
    "kola": "🥤",
}

users = {}
orders = {}

ANA_MENU = ReplyKeyboardMarkup(
    [
        ["🍔 Menü", "🛒 Sepet"],
        ["📍 Adres", "💳 Ödeme"],
        ["🔁 Tekrar Sipariş", "🗑️ Son Ürünü Sil"],
        ["✅ Onay", "❌ İptal"],
    ],
    resize_keyboard=True
)

URUN_MENU = ReplyKeyboardMarkup(
    [
        ["Pizza", "Burger"],
        ["Ayran", "Kola"],
        ["🛒 Sepet", "⬅️ Geri"],
    ],
    resize_keyboard=True
)

SEPET_MENU = ReplyKeyboardMarkup(
    [
        ["✅ Onay", "🗑️ Son Ürünü Sil"],
        ["🍔 Menü", "⬅️ Geri"],
        ["❌ İptal"],
    ],
    resize_keyboard=True
)

ODEME_MENU = ReplyKeyboardMarkup(
    [
        ["Nakit", "Kart"],
        ["⬅️ Geri", "❌ İptal"],
    ],
    resize_keyboard=True
)

ONAY_MENU = ReplyKeyboardMarkup(
    [
        ["✅ Onayla", "❌ İptal"],
        ["🛒 Sepet", "⬅️ Geri"],
    ],
    resize_keyboard=True
)


def init_excel():
    try:
        openpyxl.load_workbook(FILE_NAME)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders"
        ws.append([
            "Sipariş No",
            "Tarih",
            "Saat",
            "Müşteri",
            "Telefon",
            "Ürünler",
            "Toplam",
            "Adres",
            "Ödeme",
        ])
        wb.save(FILE_NAME)


def save_order(order_id, name, phone, cart, address, payment):
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb.active

    total = sum(MENU_ITEMS[item] for item in cart)
    now = datetime.now()

    ws.append([
        order_id,
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M"),
        name,
        phone,
        ", ".join(cart),
        total,
        address,
        payment,
    ])

    wb.save(FILE_NAME)


def is_open_now():
    hour = datetime.now().hour
    return OPEN_HOUR <= hour < CLOSE_HOUR


def make_order_id():
    return random.randint(1000, 9999)


def get_user(user_id):
    if user_id not in users:
        users[user_id] = {
            "cart": [],
            "address": None,
            "payment": None,
            "name": None,
            "phone": None,
            "waiting_name": False,
            "waiting_phone": False,
            "waiting_address": False,
            "last_order": [],
            "state": "main",
        }
    return users[user_id]


def set_state(user_id, state):
    users[user_id]["state"] = state


def reset_waiting_flags(user):
    user["waiting_name"] = False
    user["waiting_phone"] = False
    user["waiting_address"] = False


def reset_user(user_id, keep_address=True, keep_last_order=True, keep_name=True, keep_phone=True):
    old = users[user_id]
    users[user_id] = {
        "cart": [],
        "address": old["address"] if keep_address else None,
        "payment": None,
        "name": old["name"] if keep_name else None,
        "phone": old["phone"] if keep_phone else None,
        "waiting_name": False,
        "waiting_phone": False,
        "waiting_address": False,
        "last_order": old["last_order"] if keep_last_order else [],
        "state": "main",
    }


def cart_total(user):
    return sum(MENU_ITEMS[item] for item in user["cart"])


def format_cart(user):
    if not user["cart"]:
        return "Sepetin şu an boş 😕"

    lines = []
    for item in user["cart"]:
        lines.append(f"{EMOJI[item]} {item.capitalize()} - {MENU_ITEMS[item]} TL")

    address = user["address"] if user["address"] else "Girilmedi"
    payment = user["payment"].capitalize() if user["payment"] else "Seçilmedi"

    return (
        f"🛒 {RESTAURANT_NAME} Sepet Özeti\n\n"
        + "\n".join(lines)
        + f"\n\n💰 Toplam: {cart_total(user)} TL"
        + f"\n📍 Adres: {address}"
        + f"\n💳 Ödeme: {payment}"
    )


async def go_back(update: Update, context: ContextTypes.DEFAULT_TYPE, user_id: int):
    user = get_user(user_id)
    state = user["state"]

    if state == "menu":
        set_state(user_id, "main")
        await update.message.reply_text("Ana menüye döndün 👇", reply_markup=ANA_MENU)
        return

    if state == "cart":
        set_state(user_id, "main")
        await update.message.reply_text("Ana menüye döndün 👇", reply_markup=ANA_MENU)
        return

    if state == "payment":
        set_state(user_id, "cart")
        await update.message.reply_text(
            format_cart(user),
            reply_markup=SEPET_MENU
        )
        return

    if state == "confirm":
        set_state(user_id, "cart")
        await update.message.reply_text(
            format_cart(user),
            reply_markup=SEPET_MENU
        )
        return

    if state == "address":
        reset_waiting_flags(user)
        set_state(user_id, "main")
        await update.message.reply_text("Ana menüye döndün 👇", reply_markup=ANA_MENU)
        return

    await update.message.reply_text("Zaten ana menüdesin 👇", reply_markup=ANA_MENU)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    name = update.effective_user.first_name or "Dostum"
    user = get_user(user_id)

    status = (
        f"🟢 Açığız ({OPEN_HOUR}:00 - {CLOSE_HOUR}:00)"
        if is_open_now()
        else f"🔴 Şu an kapalıyız ({OPEN_HOUR}:00 - {CLOSE_HOUR}:00)"
    )

    text = (
        f"🍔 {RESTAURANT_NAME}\n\n"
        f"Hoş geldin {name} 👋\n"
        f"{status}\n\n"
        "Sipariş vermek için aşağıdaki menüyü kullanabilirsin."
    )

    reset_waiting_flags(user)
    set_state(user_id, "main")

    await update.message.reply_text(text, reply_markup=ANA_MENU)


async def handle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    raw_text = update.message.text.strip()
    text = raw_text.lower()

    user = get_user(user_id)

    if user["waiting_name"]:
        user["name"] = raw_text
        user["waiting_name"] = False
        user["waiting_phone"] = True
        set_state(user_id, "phone")
        await update.message.reply_text("📞 Telefon numaranı yaz:")
        return

    if user["waiting_phone"]:
        user["phone"] = raw_text
        user["waiting_phone"] = False
        user["waiting_address"] = True
        set_state(user_id, "address")
        await update.message.reply_text("📍 Teslimat adresini yaz:")
        return

    if user["waiting_address"]:
        user["address"] = raw_text
        user["waiting_address"] = False
        set_state(user_id, "main")
        await update.message.reply_text("📍 Adres kaydedildi ✅", reply_markup=ANA_MENU)
        return

    if "geri" in text:
        await go_back(update, context, user_id)
        return

    if "iptal" in text:
        reset_user(user_id, keep_address=True, keep_last_order=True, keep_name=True, keep_phone=True)
        await update.message.reply_text(
            "Sipariş iptal edildi ❌\nYeni sipariş için menüyü kullanabilirsin.",
            reply_markup=ANA_MENU
        )
        return

    if "menü" in text or "menu" in text:
        set_state(user_id, "menu")
        await update.message.reply_text("🍔 Ürün seçebilirsin 👇", reply_markup=URUN_MENU)
        return

    if "tekrar sipariş" in text:
        if not user["last_order"]:
            await update.message.reply_text(
                "Henüz tekrar edilecek siparişin yok 😕",
                reply_markup=ANA_MENU
            )
            return

        user["cart"] = list(user["last_order"])
        set_state(user_id, "cart")
        await update.message.reply_text(
            "🔁 Son siparişin sepete eklendi!",
            reply_markup=SEPET_MENU
        )
        return

    if "son ürünü sil" in text:
        if not user["cart"]:
            await update.message.reply_text(
                "Silinecek ürün yok. Sepetin boş 😕",
                reply_markup=ANA_MENU
            )
            return

        removed = user["cart"].pop()
        set_state(user_id, "cart" if user["cart"] else "main")
        await update.message.reply_text(
            f"🗑️ Son ürün kaldırıldı: {removed.capitalize()}",
            reply_markup=SEPET_MENU if user["cart"] else ANA_MENU
        )
        return

    if text in MENU_ITEMS:
        user["cart"].append(text)
        set_state(user_id, "menu")

        responses = [
            f"{EMOJI[text]} {text.capitalize()} sepete eklendi 😋",
            f"{EMOJI[text]} Harika seçim! {text.capitalize()} eklendi 🔥",
            f"{EMOJI[text]} Bunu da sepete attım 👌",
        ]

        await update.message.reply_text(
            random.choice(responses) + f"\n💰 Toplam: {cart_total(user)} TL",
            reply_markup=URUN_MENU
        )
        return

    if "sepet" in text:
        set_state(user_id, "cart")
        await update.message.reply_text(
            format_cart(user),
            reply_markup=SEPET_MENU
        )
        return

    if "adres" in text:
        user["waiting_address"] = True
        set_state(user_id, "address")
        await update.message.reply_text(
            "📍 Lütfen teslimat adresini yaz:",
            reply_markup=ReplyKeyboardRemove()
        )
        return

    if "ödeme" in text or "odeme" in text:
        set_state(user_id, "payment")
        await update.message.reply_text(
            "💳 Ödeme yöntemini seç:",
            reply_markup=ODEME_MENU
        )
        return

    if text in ["nakit", "kart"]:
        user["payment"] = text
        set_state(user_id, "main")
        await update.message.reply_text(
            f"💳 Ödeme seçildi: {text.capitalize()} ✅",
            reply_markup=ANA_MENU
        )
        return

    if "onay" in text and "onayla" not in text:
        set_state(user_id, "confirm")
        await update.message.reply_text(
            format_cart(user) + "\n\nSiparişi tamamlamak için aşağıdaki butonu kullan 👇",
            reply_markup=ONAY_MENU
        )
        return

    if "onayla" in text:
        if not is_open_now():
            await update.message.reply_text(
                f"🔴 Şu an kapalıyız.\nÇalışma saatleri: {OPEN_HOUR}:00 - {CLOSE_HOUR}:00",
                reply_markup=ANA_MENU
            )
            return

        if not user["cart"]:
            await update.message.reply_text(
                "Sepetin boş ❌\nÖnce ürün seçmelisin.",
                reply_markup=ANA_MENU
            )
            return

        if not user["name"]:
            user["waiting_name"] = True
            set_state(user_id, "name")
            await update.message.reply_text("👤 İsmini yaz:")
            return

        if not user["phone"]:
            user["waiting_phone"] = True
            set_state(user_id, "phone")
            await update.message.reply_text("📞 Telefon numaranı yaz:")
            return

        if not user["address"]:
            user["waiting_address"] = True
            set_state(user_id, "address")
            await update.message.reply_text(
                "Adres yazmadan siparişi tamamlayamazsın ❌\nLütfen şimdi adresini yaz.",
                reply_markup=ReplyKeyboardRemove()
            )
            return

        if not user["payment"]:
            set_state(user_id, "payment")
            await update.message.reply_text(
                "❌ Ödeme yöntemi seçmedin.\n\nLütfen aşağıdan seç 👇",
                reply_markup=ODEME_MENU
            )
            return

        total = cart_total(user)
        order_id = make_order_id()
        orders[order_id] = user_id

        keyboard = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("🍳 Hazırlanıyor", callback_data=f"prep_{order_id}"),
                InlineKeyboardButton("🚗 Yolda", callback_data=f"road_{order_id}"),
            ],
            [
                InlineKeyboardButton("✅ Teslim", callback_data=f"done_{order_id}")
            ]
        ])

        admin_text = (
            f"🚨 YENİ SİPARİŞ - {RESTAURANT_NAME}\n\n"
            f"🧾 Sipariş No: #{order_id}\n"
            f"👤 {user['name']}\n"
            f"📞 {user['phone']}\n"
            f"🛒 {', '.join(item.capitalize() for item in user['cart'])}\n"
            f"💰 {total} TL\n"
            f"📍 {user['address']}\n"
            f"💳 {user['payment'].capitalize()}\n"
            f"🕒 {datetime.now().strftime('%H:%M')}"
        )

        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=admin_text,
            reply_markup=keyboard
        )

        save_order(
            order_id,
            user["name"],
            user["phone"],
            user["cart"],
            user["address"],
            user["payment"]
        )

        user["last_order"] = list(user["cart"])

        await update.message.reply_text(
            f"🎉 Siparişin alındı {user['name']}!\n\n"
            f"🧾 Sipariş numaran: #{order_id}\n"
            f"⏱ Ortalama teslim süresi: 20-30 dk\n"
            f"📞 Gerekirse seninle iletişime geçebiliriz.\n\n"
            "Teşekkür ederiz ❤️",
            reply_markup=ANA_MENU
        )

        reset_user(user_id, keep_address=True, keep_last_order=True, keep_name=True, keep_phone=True)
        return

    await update.message.reply_text(
        "Seni tam anlayamadım 😕\n\nAşağıdaki menüden devam edebilirsin 👇",
        reply_markup=ANA_MENU
    )


async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    data = query.data
    order_id = int(data.split("_")[1])
    user_id = orders.get(order_id)

    if not user_id:
        return

    if data.startswith("prep"):
        msg = "🍳 Siparişin hazırlanıyor!"
    elif data.startswith("road"):
        msg = "🚗 Siparişin yolda!"
    else:
        msg = "✅ Sipariş teslim edildi!"

    await context.bot.send_message(chat_id=user_id, text=msg)


def main():
    init_excel()

    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle))
    app.add_handler(CallbackQueryHandler(button_handler))

    print("Bot çalışıyor...")
    app.run_polling()


if __name__ == "__main__":
    main()
